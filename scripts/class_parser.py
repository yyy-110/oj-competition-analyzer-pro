#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
class_parser.py - 班级名单智能解析模块 (Pro)
"""

import os
import re
import sys
import json

def clean_name(name):
    if name is None:
        return ""
    name_str = str(name).strip()
    name_str = re.sub(r"[\s\u3000\t\r\n]+", "", name_str)
    return name_str

def parse_excel_roster(file_path):
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("需要 openpyxl 库以解析 Excel 格式的班级名单，请先安装: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    class_to_students = {}
    student_to_class = {}

    for sheet in wb.worksheets:
        max_r = sheet.max_row
        max_c = sheet.max_column
        if not max_r or not max_c or max_r < 1 or max_c < 1:
            continue

        row1 = [sheet.cell(1, c).value for c in range(1, max_c + 1)]
        row1_clean = [clean_name(v) for v in row1]

        class_col_idx = None
        name_col_idx = None
        user_col_idx = None

        for idx, header in enumerate(row1_clean, 1):
            if any(k in header for k in ["班级", "班", "Class", "class"]):
                class_col_idx = idx
            elif any(k in header for k in ["姓名", "学生", "Name", "name", "Nick", "nick"]):
                name_col_idx = idx
            elif any(k in header for k in ["学号", "账号", "User", "user", "ID", "id"]):
                user_col_idx = idx

        if class_col_idx and name_col_idx:
            for r in range(2, max_r + 1):
                cls_val = clean_name(sheet.cell(r, class_col_idx).value)
                name_val = clean_name(sheet.cell(r, name_col_idx).value)
                user_val = clean_name(sheet.cell(r, user_col_idx).value) if user_col_idx else ""
                
                if not cls_val or not name_val:
                    continue

                if cls_val not in class_to_students:
                    class_to_students[cls_val] = []
                if name_val not in class_to_students[cls_val]:
                    class_to_students[cls_val].append(name_val)
                student_to_class[name_val] = cls_val
                if user_val:
                    student_to_class[user_val] = cls_val
        else:
            for c in range(1, max_c + 1):
                col_header = sheet.cell(1, c).value
                if col_header is None:
                    continue
                cls_name = clean_name(col_header)
                if not cls_name:
                    continue

                if cls_name not in class_to_students:
                    class_to_students[cls_name] = []

                for r in range(2, max_r + 1):
                    val = sheet.cell(r, c).value
                    if val is None:
                        continue
                    name_clean = clean_name(val)
                    if name_clean and name_clean not in class_to_students[cls_name]:
                        class_to_students[cls_name].append(name_clean)
                        student_to_class[name_clean] = cls_name

    wb.close()
    return class_to_students, student_to_class

def parse_class_roster(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"班级名单文件不存在: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".xlsx", ".xls", ".xlsm"]:
        return parse_excel_roster(file_path)
    elif ext in [".csv", ".txt"]:
        class_to_students = {}
        student_to_class = {}
        with open(file_path, "r", encoding="utf-8-sig", errors="ignore") as f:
            lines = [l.strip() for l in f if l.strip()]
        if not lines:
            return class_to_students, student_to_class
        
        delimiter = "\t" if "\t" in lines[0] else ","
        headers = [clean_name(x) for x in lines[0].split(delimiter)]
        
        for c, cls_name in enumerate(headers):
            if not cls_name:
                continue
            if cls_name not in class_to_students:
                class_to_students[cls_name] = []
            for line in lines[1:]:
                parts = [clean_name(x) for x in line.split(delimiter)]
                if c < len(parts) and parts[c]:
                    st_name = parts[c]
                    if st_name not in class_to_students[cls_name]:
                        class_to_students[cls_name].append(st_name)
                        student_to_class[st_name] = cls_name
        return class_to_students, student_to_class
    elif ext == ".json":
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            class_to_students = {clean_name(k): [clean_name(s) for s in v if clean_name(s)] for k, v in data.items()}
            student_to_class = {}
            for cls_name, sts in class_to_students.items():
                for st in sts:
                    student_to_class[st] = cls_name
            return class_to_students, student_to_class
        else:
            raise ValueError("JSON 格式班级名单需为 { '班级名': ['学生1', '学生2'] } 结构")
    else:
        raise ValueError(f"不支持的班级名单文件类型: {ext} (推荐使用 .xlsx 格式)")

def match_user_to_class(user, student_to_class):
    clean_u = clean_name(user)

    if clean_u and clean_u in student_to_class:
        return student_to_class[clean_u]

    for st_name, cls_name in student_to_class.items():
        if st_name and len(st_name) >= 2:
            if clean_u and (st_name == clean_u or st_name in clean_u or clean_u in st_name):
                return cls_name

    return "未分班"

def build_class_mapping(rank_users, all_subs, roster_file=None):
    all_users = list(set([r["user"] for r in rank_users] + [s["user"] for s in all_subs]))
    
    if not roster_file or not os.path.exists(roster_file):
        return {
            "has_roster": False,
            "user_class_map": {u: "全年级" for u in all_users},
            "name_or_user_to_class": {u: "全年级" for u in all_users},
            "class_students_map": {"全年级": all_users},
            "classes_list": ["全年级"],
            "roster_class_to_students": {},
            "unmatched_users": []
        }

    class_to_roster, student_to_class = parse_class_roster(roster_file)

    user_class_map = {}
    name_or_user_to_class = dict(student_to_class)
    class_students_map = {cls: [] for cls in class_to_roster.keys()}
    class_students_map["未分班"] = []
    unmatched_users = []

    for u in all_users:
        cls = match_user_to_class(u, student_to_class)
        user_class_map[u] = cls
        name_or_user_to_class[u] = cls

        if cls in class_students_map:
            class_students_map[cls].append(u)
        else:
            class_students_map[cls] = [u]

        if cls == "未分班":
            unmatched_users.append({"user": u, "nick": u})

    classes_list = [c for c in class_to_roster.keys() if len(class_students_map.get(c, [])) > 0 or len(class_to_roster.get(c, [])) > 0]
    if len(class_students_map.get("未分班", [])) > 0:
        classes_list.append("未分班")

    return {
        "has_roster": True,
        "user_class_map": user_class_map,
        "name_or_user_to_class": name_or_user_to_class,
        "class_students_map": class_students_map,
        "classes_list": classes_list,
        "roster_class_to_students": class_to_roster,
        "unmatched_users": unmatched_users
    }
